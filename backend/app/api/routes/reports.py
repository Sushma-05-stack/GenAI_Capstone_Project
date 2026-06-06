"""
/reports/export - PDF, CSV, Excel report generation
"""
import io
import csv
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import Optional
from app.api.deps import get_current_user
from app.models.user import User
from app.models.evaluation import EvaluationRun, EvaluationResult, EvalStatus
from app.security.audit import log_event
from app.models.audit import AuditAction

router = APIRouter(prefix="/reports", tags=["Reports"])


@router.get("/export/csv")
async def export_csv(
    run_id: str,
    current_user: User = Depends(get_current_user),
):
    """Export evaluation results as CSV."""
    run = await EvaluationRun.get(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")

    results = await EvaluationResult.find(EvaluationResult.run_id == run_id).to_list()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Question", "Answer", "Ground Truth",
        "Faithfulness", "Answer Relevancy", "Context Precision",
        "Context Recall", "Hallucination Risk", "Retrieval Quality",
        "Latency (ms)", "Cost (USD)", "Model", "Provider", "Fallback Used"
    ])
    for r in results:
        writer.writerow([
            r.question, r.answer, r.ground_truth or "",
            r.faithfulness, r.answer_relevancy, r.context_precision,
            r.context_recall, r.hallucination_risk, r.retrieval_quality,
            r.latency_ms, r.cost_usd, r.model_used, r.provider_used, r.fallback_used,
        ])

    output.seek(0)
    await log_event(action=AuditAction.EXPORT_REPORT, user_id=str(current_user.id), resource=run_id)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=eval_run_{run_id}.csv"},
    )


@router.get("/export/excel")
async def export_excel(
    run_id: str,
    current_user: User = Depends(get_current_user),
):
    """Export evaluation results as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    run = await EvaluationRun.get(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")

    results = await EvaluationResult.find(EvaluationResult.run_id == run_id).to_list()

    wb = openpyxl.Workbook()
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    summary_rows = [
        ("Run Name", run.name),
        ("Model", run.model_name),
        ("Provider", run.provider),
        ("Total Questions", run.total_questions),
        ("Avg Faithfulness", run.avg_faithfulness),
        ("Avg Answer Relevancy", run.avg_answer_relevancy),
        ("Avg Context Precision", run.avg_context_precision),
        ("Avg Context Recall", run.avg_context_recall),
        ("Avg Hallucination Risk", run.avg_hallucination_risk),
        ("Avg Latency (ms)", run.avg_latency_ms),
        ("Total Cost (USD)", run.total_cost_usd),
    ]
    for row in summary_rows:
        ws_summary.append(row)

    # Results sheet
    ws_results = wb.create_sheet("Results")
    headers = [
        "Question", "Answer", "Ground Truth",
        "Faithfulness", "Answer Relevancy", "Context Precision",
        "Context Recall", "Hallucination Risk", "Retrieval Quality",
        "Latency (ms)", "Cost (USD)", "Model", "Provider", "Fallback"
    ]
    ws_results.append(headers)
    for cell in ws_results[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r in results:
        ws_results.append([
            r.question[:200], r.answer[:200], r.ground_truth or "",
            r.faithfulness, r.answer_relevancy, r.context_precision,
            r.context_recall, r.hallucination_risk, r.retrieval_quality,
            r.latency_ms, r.cost_usd, r.model_used, r.provider_used, r.fallback_used,
        ])

    # Auto-width
    for col in ws_results.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_results.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    await log_event(action=AuditAction.EXPORT_REPORT, user_id=str(current_user.id), resource=run_id)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=eval_run_{run_id}.xlsx"},
    )


@router.get("/export/pdf")
async def export_pdf(
    run_id: str,
    current_user: User = Depends(get_current_user),
):
    """Generate a summary PDF report."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    run = await EvaluationRun.get(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    # Title
    story.append(Paragraph(f"RAG Evaluation Report: {run.name}", styles["Title"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"Model: {run.provider}/{run.model_name}", styles["Normal"]))
    story.append(Paragraph(f"Status: {run.status}", styles["Normal"]))
    story.append(Spacer(1, 20))

    # Metrics table
    metrics_data = [
        ["Metric", "Score"],
        ["Faithfulness", str(run.avg_faithfulness or "N/A")],
        ["Answer Relevancy", str(run.avg_answer_relevancy or "N/A")],
        ["Context Precision", str(run.avg_context_precision or "N/A")],
        ["Context Recall", str(run.avg_context_recall or "N/A")],
        ["Hallucination Risk", str(run.avg_hallucination_risk or "N/A")],
        ["Avg Latency (ms)", str(run.avg_latency_ms or "N/A")],
        ["Total Cost (USD)", str(run.total_cost_usd or "N/A")],
    ]
    table = Table(metrics_data, colWidths=[250, 150])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4FF")]),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
    ]))
    story.append(table)
    doc.build(story)

    buffer.seek(0)
    await log_event(action=AuditAction.EXPORT_REPORT, user_id=str(current_user.id), resource=run_id)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=eval_report_{run_id}.pdf"},
    )
